from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from config import ARQUIVO_EXCEL

def criar_relatorio_excel(dados_moedas, taxa_selic, data_execucao, historico_dolar):
    """
    Gera um arquivo Excel com múltiplas abas contendo dados de moedas, taxas de juros,
    histórico do dólar e log de execução.

    A planilha contém as seguintes abas:
    - "Resumo Atual": valores atuais das moedas (USD, EUR, BTC).
    - "Taxas de Juros": histórico mensal da taxa SELIC dos últimos 6 meses.
    - "Log de Execução": data e hora da geração do relatório.
    - "Histórico - Dólar": valores diários do dólar nos últimos 30 dias.

    Args:
        dados_moedas (dict): Dicionário com os nomes das moedas e seus respectivos valores em BRL.
        taxa_selic (dict): Dicionário no formato {"MM/YYYY": valor}, com a taxa SELIC mensal.
        data_execucao (datetime): Data e hora em que o relatório está sendo gerado.
        historico_dolar (dict): Dicionário com datas ("YYYY-MM-DD") e valores da cotação do dólar.

    Returns:
        None: O arquivo Excel é salvo diretamente no caminho definido por ARQUIVO_EXCEL.
    """
    
    wb = Workbook()

    # Aba 1: Resumo Atual
    aba_resumo = wb.active
    aba_resumo.title = "Resumo Atual"
    aba_resumo.append(["Moeda", "Valor (R$)"])
    for cell in aba_resumo[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for moeda, valor in dados_moedas.items():
        aba_resumo.append([moeda, round(valor, 2)])
        aba_resumo[f"B{aba_resumo.max_row}"].number_format = '0.00'

    # Aba 2: Taxas de Juros
    aba_juros = wb.create_sheet("Taxas de Juros")
    aba_juros.append(["Data", "SELIC (%)"])
    for cell in aba_juros[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for data, valor in taxa_selic.items():
        aba_juros.append([data, round(valor, 2)])
    for row in aba_juros.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.00'

    # Aba 3: Log de Execução
    aba_log = wb.create_sheet("Log de Execução")
    aba_log.append(["Data/Hora da Coleta"])
    aba_log.append([data_execucao.strftime("%Y-%m-%d %H:%M:%S")])
    for cell in aba_log[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Aba 4: Histórico - Dólar
    aba_hist = wb.create_sheet("Histórico - Dólar")
    aba_hist.append(["Data", "Valor (R$)"])
    for cell in aba_hist[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for data, valor in historico_dolar.items():
        aba_hist.append([data, round(valor, 2)])

    for row in aba_hist.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.00'

    wb.save(ARQUIVO_EXCEL)